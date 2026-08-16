"""PaymentService - Nghiệp vụ bảng lương, cột động và export."""
from datetime import datetime
from decimal import Decimal

from django.utils import timezone
from rest_framework.exceptions import ValidationError

from apps.duties.models import (
    DutySchedule,
    InstructorAttendanceLog,
    InstructorPayment,
    InstructorPaymentColumn,
    InstructorPaymentDetail,
)
from apps.duties.repositories import duty_repository, payment_repository
from apps.duties.services.config import _cfg

# Giới hạn số cột động (Thưởng/Khấu trừ) tối đa khi export
MAX_EXPORT_DYNAMIC_COLUMNS = 30


def compute_all_payments(month):
    """Tổng hợp lương tự động cho TẤT CẢ giảng viên trong 1 tháng (YYYY-MM)."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    instructors = User.objects.filter(role__code="INSTRUCTOR", is_active=True)
    results = []
    for inst in instructors:
        try:
            p = compute_monthly_payment(inst.id, month)
            results.append({"instructor_id": inst.id, "status": p.status, "net": float(p.net_amount)})
        except Exception:
            continue
    return results


def approve_payment(user, payment_id, approve=True):
    """Admin hoặc hủy duyệt bảng lương."""
    payment = payment_repository.get_payment_by_id(payment_id)
    if not payment:
        raise ValidationError("Không tìm thấy bảng lương.")
    if approve:
        if payment.status != InstructorPayment.Status.DRAFT:
            raise ValidationError("Chỉ bảng lương DRAFT mới duyệt được.")
        payment.status = InstructorPayment.Status.APPROVED
        payment.approved_by = user
    else:
        if payment.status == InstructorPayment.Status.PAID:
            raise ValidationError("Bảng lương đã chi trả không thể hủy duyệt.")
        payment.status = InstructorPayment.Status.DRAFT
        payment.approved_by = None
    payment.save()
    return payment


def mark_payment_paid(user, payment_id):
    """Admin xác nhận thanh toán: thực sự chuyển tiền qua Stripe rồi đánh dấu PAID.

    - Lấy Stripe Connected Account (acct_...) từ bank_account_number của InstructorProfile.
    - Gọi transfer_to_instructor để chuyển net_amount từ platform sang tài khoản giảng viên.
    - Nếu chuyển tiền thành công → đổi status APPROVED→PAID + lưu paid_at.
    - Nếu thiếu Connected Account hoặc chuyển thất bại → báo lỗi, KHÔNG đánh dấu PAID.
    """
    from apps.payments.services.stripe_payment_service import transfer_to_instructor

    payment = payment_repository.get_payment_by_id(payment_id)
    if not payment:
        raise ValidationError("Không tìm thấy bảng lương.")
    if payment.status != InstructorPayment.Status.APPROVED:
        raise ValidationError("Chỉ bảng lương APPROVED mới đánh dấu đã chi trả.")

    instructor = payment.instructor
    profile = getattr(instructor, "instructor_profile", None)
    destination_account = None
    if profile is not None and profile.bank_account_number:
        acct = str(profile.bank_account_number).strip()
        if acct.startswith("acct_"):
            destination_account = acct

    if not destination_account:
        raise ValidationError(
            "Giảng viên chưa có Stripe Connected Account (acct_...). "
            "Vui lòng cập nhật thông tin ngân hàng trước khi thanh toán."
        )

    # Chuyển tiền thật trên Stripe: platform -> connected account của giảng viên
    amount = int(payment.net_amount)
    transfer = transfer_to_instructor(
        destination_account,
        amount,
        metadata={
            "type": "instructor_payroll",
            "payment_id": str(payment.id),
            "instructor_id": str(instructor.id),
            "month": payment.month,
            "confirmed_by": user.email if user else "",
        },
    )

    # Lưu transfer id để truy vết
    payment.provider_transfer_id = getattr(transfer, "id", None) or getattr(transfer, "sid", None)
    payment.status = InstructorPayment.Status.PAID
    payment.paid_at = timezone.now()
    payment.save()
    return payment


# ==================== CỘT ĐỘNG BẢNG LƯƠNG ====================

def list_payment_columns():
    """Danh sách cột động toàn cục (Thưởng/Khấu trừ)."""
    return payment_repository.list_payment_columns()


def create_payment_column(name, column_type):
    """Tạo cột động toàn cục."""
    if not name or not str(name).strip():
        raise ValidationError("Vui lòng nhập tên cột.")
    column_type = (column_type or "").upper()
    if column_type not in (InstructorPaymentColumn.Type.BONUS, InstructorPaymentColumn.Type.DEDUCTION):
        raise ValidationError("Loại cột phải là BONUS hoặc DEDUCTION.")
    if payment_repository.get_payment_column_by_name(str(name).strip()):
        raise ValidationError("Cột này đã tồn tại.")
    if payment_repository.count_payment_columns() >= MAX_EXPORT_DYNAMIC_COLUMNS:
        raise ValidationError(
            f"Số cột động đã đạt giới hạn tối đa {MAX_EXPORT_DYNAMIC_COLUMNS} cột. Không thể tạo thêm cột."
        )
    return payment_repository.create_payment_column(str(name).strip(), column_type)


def delete_payment_column(column_id):
    """Xóa cột động toàn cục cùng mọi giá trị + tính lại thực nhận cho các bảng lương liên quan."""
    col = payment_repository.get_payment_column_by_id(column_id)
    if not col:
        raise ValidationError("Không tìm thấy cột.")
    payment_ids = payment_repository.get_payment_ids_having_column(col)
    payment_repository.delete_payment_column(col)
    for pid in payment_ids:
        p = payment_repository.get_payment_by_id(pid)
        if p:
            _recalc_net_from_columns(p)
    return col


def set_payment_column_value(user, payment_id, column_id, amount):
    """Nhập giá trị của cột động cho 1 bảng lương (Thưởng cộng / Khấu trừ trừ)."""
    payment = payment_repository.get_payment_by_id(payment_id)
    if not payment:
        raise ValidationError("Không tìm thấy bảng lương.")
    if payment.status == InstructorPayment.Status.PAID:
        raise ValidationError("Bảng lương đã chi trả không thể sửa.")
    col = payment_repository.get_payment_column_by_id(column_id)
    if not col:
        raise ValidationError("Không tìm thấy cột.")
    try:
        a = Decimal(str(amount))
    except Exception:
        raise ValidationError("Giá trị phải là một số.")
    payment_repository.update_or_create_column_value(
        payment, col, defaults={"amount": a, "updated_at": timezone.now()},
    )
    return _recalc_net_from_columns(payment)


def _recalc_net_from_columns(payment):
    """Tính lại net_amount từ cột động (BONUS/DEDUCTION).

    Thực nhận = Lương (regular + overtime) + tổng Thưởng − tổng Khấu trừ (cột động).
    """
    bonus_total = Decimal("0")
    deduction_cols_total = Decimal("0")

    for col in payment_repository.list_payment_columns():
        v = payment_repository.get_column_value(payment, col)
        if not v:
            continue
        if col.column_type == InstructorPaymentColumn.Type.BONUS:
            bonus_total += v.amount
        else:
            deduction_cols_total += v.amount

    payment.net_amount = (
        payment.salary_amount
        + bonus_total
        - deduction_cols_total
    )
    payment.save(update_fields=["net_amount", "updated_at"])
    return payment


# ==================== LƯƠNG ====================

def compute_monthly_payment(instructor_id, month):
    """Tính lương tháng - tổng hợp TẤT CẢ ca chấm công hợp lệ.

    Lương = regular_hours * lương tối thiểu + overtime_hours * lương giờ thêm.
    Thực nhận = Lương + Thưởng − Khấu trừ (cột động).
    """
    base_rate = Decimal(str(_cfg("duty_salary_min_rate", 200000)))
    ot_rate = Decimal(str(_cfg("duty_salary_overtime_rate", 220000)))
    min_hours = Decimal(str(_cfg("duty_min_teaching_hours", 30)))

    logs = duty_repository.list_attendance_logs_for_payment(
        instructor_id,
        int(month[:4]),
        int(month[5:7]),
        [
            InstructorAttendanceLog.Status.OK,
            InstructorAttendanceLog.Status.LATE,
            InstructorAttendanceLog.Status.EARLY_LEAVE,
            InstructorAttendanceLog.Status.LATE_EARLY,
        ],
    )

    total_minutes = duty_repository.sum_counted_minutes(
        instructor_id,
        year=int(month[:4]),
        month=int(month[5:7]),
        statuses=[
            InstructorAttendanceLog.Status.OK,
            InstructorAttendanceLog.Status.LATE,
            InstructorAttendanceLog.Status.EARLY_LEAVE,
            InstructorAttendanceLog.Status.LATE_EARLY,
        ],
    )
    total_hours = (Decimal(str(total_minutes)) / Decimal("60")).quantize(Decimal("0.01"))
    regular_hours = min(total_hours, min_hours)
    overtime_hours = max(Decimal("0"), total_hours - min_hours)

    regular_amount = (regular_hours * base_rate).quantize(Decimal("0.01"))
    overtime_amount = (overtime_hours * ot_rate).quantize(Decimal("0.01"))
    salary_amount = regular_amount + overtime_amount
    net_amount = salary_amount

    payment, _ = payment_repository.update_or_create_payment(
        instructor_id, month,
        defaults={
            "regular_hours": regular_hours, "overtime_hours": overtime_hours,
            "total_hours": total_hours, "regular_amount": regular_amount,
            "overtime_amount": overtime_amount, "salary_amount": salary_amount,
            "net_amount": net_amount,
            "reviewed_by": None, "approved_by": None,
        },
    )

    InstructorPaymentDetail.objects.filter(payment=payment).delete()
    for log in logs:
        InstructorPaymentDetail.objects.create(
            payment=payment, schedule=log.schedule, attendance=log,
            date=log.login_at.date(), shift=log.schedule.shift if log.schedule else "-",
            actual_minutes=log.actual_minutes, counted_minutes=log.counted_minutes,
            missing_minutes=log.missing_minutes,
            paid_amount=(Decimal(str(log.counted_minutes)) / Decimal("60") * base_rate).quantize(Decimal("0.01")),
        )

    # Cập nhật lại thực nhận sau khi có cột động (nếu bảng lương này có Thưởng/Phạt)
    payment = _recalc_net_from_columns(payment)
    return payment


def list_instructor_missing_hours(instructor_id, month=None):
    """Danh sách các ca thiếu giờ chưa có ca bù - phục vụ sắp lịch bù."""
    from django.db.models import Exists, OuterRef

    grace_minutes = _cfg("duty_grace_minutes", 15)
    logs = duty_repository.list_missing_hours_logs(
        instructor_id, month=month, min_missing_minutes=grace_minutes,
    )
    makeup_exists = DutySchedule.objects.filter(
        makeup_for=OuterRef("schedule"),
        status__in=[DutySchedule.Status.SCHEDULED, DutySchedule.Status.DONE],
    )
    return logs.annotate(
        has_makeup=Exists(makeup_exists)
    ).filter(has_makeup=False).order_by("login_at")


# ==================== EXPORT EXCEL ====================

def list_payments_for_export(month, instructor_id=None):
    """Lấy danh sách bảng lương theo tháng (bắt buộc) + tùy chọn lọc giảng viên.

    - Bắt buộc có tháng + năm (định dạng YYYY-MM).
    - Chỉ xuất các bảng lương ở trạng thái APPROVED (Đã duyệt).
    - Nếu có instructor_id: chỉ xuất bảng lương của 1 giảng viên đó.
    - Nếu không: xuất tất cả giảng viên trong tháng đó.
    """
    if not month:
        raise ValidationError("Thiếu tháng và năm. Vui lòng chọn tháng/năm trước khi xuất file.")
    try:
        datetime.strptime(str(month), "%Y-%m")
    except Exception:
        raise ValidationError("Tháng không hợp lệ. Định dạng phải là YYYY-MM.")

    return payment_repository.list_payments_by_month(
        month=str(month),
        status=InstructorPayment.Status.APPROVED,
        instructor_id=instructor_id,
    )


def export_payments_excel(month, instructor_id=None):
    """Xuất file Excel (.xlsx) bảng lương theo tháng + năm đã chọn."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    payments = list_payments_for_export(month, instructor_id)
    if not payments:
        if instructor_id:
            raise ValidationError(
                f"Không có bảng lương ĐÃ DUYỆT của giảng viên này trong tháng {month}. "
                "Chỉ bảng lương ở trạng thái Đã duyệt mới được xuất file."
            )
        raise ValidationError(
            f"Không có bảng lương ĐÃ DUYỆT nào trong tháng {month}. "
            "Chỉ bảng lương ở trạng thái Đã duyệt mới được xuất file."
        )

    columns = list(payment_repository.list_payment_columns())
    if len(columns) > MAX_EXPORT_DYNAMIC_COLUMNS:
        raise ValidationError(
            f"Số cột động ({len(columns)}) vượt quá giới hạn tối đa {MAX_EXPORT_DYNAMIC_COLUMNS} cột khi export. "
            "Vui lòng xóa bớt cột Thưởng/Khấu trừ trước khi xuất file."
        )

    payment_ids = [p.id for p in payments]
    value_map = {}
    if payment_ids:
        for v in payment_repository.list_column_values_by_payments(payment_ids):
            value_map[(str(v.payment_id), str(v.column_id))] = float(v.amount)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Bảng lương {month}"

    headers = [
        "STT", "Giảng viên", "Lương tháng", "Ngày kết toán",
        "Giờ chuẩn", "Giờ thêm", "Tổng giờ làm",
        "Lương giờ chuẩn", "Lương giờ thêm", "Thành tiền",
    ]
    for c in columns:
        headers.append(f"{c.name} ({'+' if c.column_type == InstructorPaymentColumn.Type.BONUS else '−'})")
    headers += ["Thực nhận"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    money_fmt = '#,##0"đ"'
    num_fmt = "0.00"
    for row_idx, p in enumerate(payments, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=p.instructor.get_full_name() or p.instructor.email)
        ws.cell(row=row_idx, column=3, value=p.month)
        ws.cell(row=row_idx, column=4, value=p.updated_at.date().isoformat() if p.updated_at else "")
        ws.cell(row=row_idx, column=5, value=float(p.regular_hours)).number_format = num_fmt
        ws.cell(row=row_idx, column=6, value=float(p.overtime_hours)).number_format = num_fmt
        ws.cell(row=row_idx, column=7, value=float(p.total_hours)).number_format = num_fmt
        ws.cell(row=row_idx, column=8, value=float(p.regular_amount)).number_format = money_fmt
        ws.cell(row=row_idx, column=9, value=float(p.overtime_amount)).number_format = money_fmt
        ws.cell(row=row_idx, column=10, value=float(p.salary_amount)).number_format = money_fmt

        col_pos = 11
        for c in columns:
            cell = ws.cell(row=row_idx, column=col_pos, value=value_map.get((str(p.id), str(c.id)), 0))
            cell.number_format = money_fmt
            col_pos += 1

        ws.cell(row=row_idx, column=col_pos, value=float(p.net_amount)).number_format = money_fmt

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx in (2,):
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in (3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    widths = [6, 28, 14, 14, 10, 10, 10, 16, 16, 16]
    for _c in columns:
        widths.append(16)
    widths += [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio